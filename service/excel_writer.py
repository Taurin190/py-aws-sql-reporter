import openpyxl
import re
import datetime
import config.format as f


class ExcelWriter:
    def __init__(self, config=None):
        self.tmp_path = "./tmp"
        if config and "tmp_path" in config.keys():
            self.tmp_path = config["tmp_path"]

    @staticmethod
    def create_sheet(name):
        wb = openpyxl.Workbook()
        if not name.endswith(".xlsx"):
            name += ".xlsx"
        wb.save(name)

    @staticmethod
    def create_sheet_with_result(name, header, result):
        wb = openpyxl.Workbook()
        sheet_names = wb.get_sheet_names()
        ws = wb.get_sheet_by_name(sheet_names[0])
        if not name.endswith(".xlsx"):
            name += ".xlsx"
        for i in range(len(header)):
            ws.cell(column=i+1, row=1).value = header[i][0]
        for i in range(len(result)):
            for j in range(len(result[0])):
                ws.cell(column=j+1, row=i+2).value = ExcelWriter._illegal_char_remover(result[i][j])
        wb.save(name)

    @staticmethod
    def _illegal_char_remover(data):
        ILLEGAL_CHARACTERS_RE = re.compile(
            r'[\000-\010]|[\013-\014]|[\016-\037]|[\x00-\x1f\x7f-\x9f]|[\uffff]')
        """Remove ILLEGAL CHARACTER."""
        if not data:
            return data
        if type(data) is str:
            return ILLEGAL_CHARACTERS_RE.sub("", str(data))
        elif type(data) is bytes:
            return int.from_bytes(data, 'big')
        elif type(data) is datetime.datetime:
            return data.strftime(f.DATE_FORMAT)
        return data
